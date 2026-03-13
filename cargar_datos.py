"""
cargar_datos.py
Importa los datos del Excel Horarios2026 a Supabase.

Uso:
  pip install supabase pandas openpyxl
  python cargar_datos.py
"""

import os
import re
import pandas as pd
from openpyxl import load_workbook
from supabase import create_client

# ── Configurá estos valores ──────────────────────────────────
SUPABASE_URL = "https://cknifwdwtenhfwfvkwyw.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImNrbmlmd2R3dGVuaGZ3ZnZrd3l3Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3MzQxNTg1MywiZXhwIjoyMDg4OTkxODUzfQ.PSSkUdIf46V1_UXiExe5duxrqhGbWWw9gEp8CxHDoSk"   # usá la service_role key para carga inicial
EXCEL_PATH   = "Horarios2026.xlsx"     # path al archivo Excel
# ─────────────────────────────────────────────────────────────

def parse_sede_cohorte(name):
    locations = {
        'INSPT': 'INSPT', 'UTE': 'UTE', 'La Matanza': 'La Matanza',
        'Parana': 'Paraná', 'Mar del Plata': 'Mar del Plata', 'FECEA': 'FECEA',
        'Concepcion Uruguay': 'Concepción del Uruguay',
        'Concepcion del Uruguay': 'Concepción del Uruguay',
        'Tucuman': 'Tucumán', 'San Miguel': 'San Miguel', 'Catamarca': 'Catamarca',
        'Santa Maria': 'Santa María', 'Bariloche': 'Bariloche',
        'Misiones': 'Misiones', 'Jujuy': 'Jujuy',
    }
    for key, display in locations.items():
        if name.startswith(key):
            return display, name[len(key):].strip()
    return name, ''

col_groups = [
    ('INSPT Agosto 24 PDS A',3,4,5),('INSPT Agosto 24 PDS B',6,7,8),
    ('INSPT Marzo 25 PDS A',9,10,11),('INSPT Marzo 25 PDS B',12,13,14),
    ('INSPT Marzo 25 PDS C',15,16,17),('INSPT Marzo 25 PDS D',18,19,20),
    ('INSPT Agosto 25 PDS',None,22,23),('INSPT Marzo 26 PDS A',24,25,26),
    ('INSPT Marzo 26 PDS B',27,28,29),('INSPT Marzo 26 PDS C',None,31,32),
    ('UTE Octubre 24 PDS',33,34,35),('UTE Abril 25 PDS',36,37,38),
    ('UTE Noviembre 25 PDS A',39,40,41),('UTE Noviembre 25 PDS B',None,43,44),
    ('La Matanza Sep 24 PDS A',45,46,47),('La Matanza Sep 24 PDS B',None,49,50),
    ('Parana Agosto 24 PDS',None,52,53),('Parana Marzo 25 PDS',None,55,56),
    ('Parana Agosto 25 PDS A',None,58,59),('Parana Agosto 25 PDS B',None,61,62),
    ('Mar del Plata Marzo 25 PDS A',None,64,65),('Mar del Plata Marzo 25 PDS B',None,67,68),
    ('FECEA Mayo 25 PDS A',69,70,71),('FECEA Mayo 25 PDS B',None,73,74),
    ('Concepcion Uruguay Marzo 25 PDS A',None,76,77),('Concepcion Uruguay Marzo 25 PDS B',None,79,80),
    ('Tucuman Sep 24 PDS',81,82,83),('Tucuman Junio 25 PDS A',84,85,86),
    ('Tucuman Junio 25 PDS B',87,88,89),('San Miguel Mayo 25 PDS',None,97,98),
    ('Catamarca Oct 24 PDS',None,103,104),('Catamarca Mayo 25 PDS',105,106,107),
    ('Santa Maria Sep 25 PDS',None,109,110),('Bariloche Agosto 24 PDS',None,118,119),
    ('Bariloche Agosto 25 PDS',None,121,122),('Misiones Mayo 25 PDS A',None,127,128),
    ('Misiones Mayo 25 PDS B',129,130,131),('Jujuy Sep 24 PDS',None,136,137),
    ('Jujuy Abril 25 PDS A',None,139,140),('Jujuy Abril 25 PDS B',None,142,143),
]

def is_header(v):
    if not v: return False
    s = str(v).strip()
    return any(k in s for k in ['INSPT','UTE','La Matanza','Parana','Mar del Plata',
        'FECEA','Concepcion','Tucuman','San Miguel','Catamarca','Santa Maria',
        'Bariloche','Misiones','Jujuy','Asignatura','Profesor','semana','horario','Mes:'])

def extract_records(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    month_rows = [2,24,49,70,94,114,136,161,182,207,232]
    month_names = {}
    for mr in month_rows:
        v = ws.cell(mr,1).value
        if v:
            match = re.search(r'([\w]+\s+\d{4})', str(v))
            month_names[mr] = match.group(1).strip() if match else str(v).strip()

    data_ranges = []
    for i, mr in enumerate(month_rows):
        start = mr + 4
        end = month_rows[i+1]-1 if i+1 < len(month_rows) else ws.max_row
        data_ranges.append((month_names[mr], start, end))

    records = []
    for mes, start_row, end_row in data_ranges:
        cur_fecha = cur_hora = None
        for row in range(start_row, end_row+1):
            c1,c2 = ws.cell(row,1).value, ws.cell(row,2).value
            if is_header(c1) or is_header(c2): continue
            if c1 and str(c1).strip(): cur_fecha = str(c1).strip()
            if c2 and str(c2).strip(): cur_hora  = str(c2).strip()
            if not cur_fecha or not cur_hora: continue
            for (sc, cc, ac, pc) in col_groups:
                asig = ws.cell(row,ac).value
                prof = ws.cell(row,pc).value
                code = ws.cell(row,cc).value if cc else None
                if not asig or is_header(asig): continue
                asig_str = str(asig).strip()
                if not asig_str or asig_str == 'Asignatura': continue
                sede, cohorte = parse_sede_cohorte(sc)
                records.append({
                    'mes': mes, 'fecha': cur_fecha, 'horario': cur_hora,
                    'sede': sede, 'cohorte': cohorte,
                    'cuatrimestre_clase': str(code).strip() if code else '',
                    'materia': asig_str,
                    'profesor': str(prof).strip() if prof else '',
                })
    return records

if __name__ == '__main__':
    print("Extrayendo datos del Excel...")
    records = extract_records(EXCEL_PATH)
    print(f"  {len(records)} registros extraídos.")

    print("Conectando a Supabase...")
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    print("Insertando en lotes...")
    batch = 50
    for i in range(0, len(records), batch):
        chunk = records[i:i+batch]
        result = sb.table('horarios').insert(chunk).execute()
        print(f"  Insertados {min(i+batch, len(records))}/{len(records)}...")

    print("✅ Listo! Todos los datos cargados.")
